import os
from datetime import datetime, date
from typing import Optional, List, Dict, Any
import json
import mimetypes
import csv

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from database import db, create_document
from schemas import SalahTime, Announcement, Asset

app = FastAPI(title="Masjid Display Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Ensure uploads directory exists and is served statically
UPLOAD_DIR = os.path.join(os.getcwd(), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Fallback data directory for when DB is unavailable
DATA_DIR = os.path.join(os.getcwd(), "data")
os.makedirs(DATA_DIR, exist_ok=True)
SALAHS_FILE = os.path.join(DATA_DIR, "salah.json")
ANN_FILE = os.path.join(DATA_DIR, "announcements.json")


def _read_json(path: str) -> Any:
    try:
        if not os.path.exists(path):
            return None
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _write_json(path: str, data: Any) -> None:
    tmp = f"{path}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, path)


@app.get("/")
def read_root():
    return {"message": "Masjid Display Backend Running"}


@app.get("/api/hello")
def hello():
    return {"message": "Hello from the backend API!"}


@app.get("/test")
def test_database():
    """Test endpoint to check if database is available and accessible"""
    response = {
        "backend": "✅ Running",
        "database": "❌ Not Available",
        "database_url": None,
        "database_name": None,
        "connection_status": "Not Connected",
        "collections": []
    }

    try:
        if db is not None:
            response["database"] = "✅ Available"
            response["database_url"] = "✅ Configured"
            response["database_name"] = db.name if hasattr(db, 'name') else "✅ Connected"
            response["connection_status"] = "Connected"
            try:
                collections = db.list_collection_names()
                response["collections"] = collections[:10]
                response["database"] = "✅ Connected & Working"
            except Exception as e:
                response["database"] = f"⚠️  Connected but Error: {str(e)[:50]}"
        else:
            response["database"] = "⚠️  Available but not initialized"

    except Exception as e:
        response["database"] = f"❌ Error: {str(e)[:50]}"

    response["database_url"] = "✅ Set" if os.getenv("DATABASE_URL") else "❌ Not Set"
    response["database_name"] = "✅ Set" if os.getenv("DATABASE_NAME") else "❌ Not Set"

    return response


# ---------------- Salah Times Endpoints ----------------

@app.get("/api/salah/today")
def get_today_salah():
    today = date.today().isoformat()
    if db is None:
        # Fallback to local JSON store
        items = _read_json(SALAHS_FILE) or {}
        return items.get(today, {"date": today})
    doc = db["salahtime"].find_one({"date": today}, {"_id": 0})
    return doc or {"date": today}


@app.get("/api/salah")
def get_salah_by_date(d: Optional[str] = None):
    if db is None:
        # Fallback to local JSON store
        store = _read_json(SALAHS_FILE) or {}
        if d is None:
            # return recent up to 30 by date desc
            keys = sorted(store.keys(), reverse=True)[:30]
            return [store[k] for k in keys]
        return store.get(d, {"date": d})
    if d is None:
        # return recent 30 entries
        cur = db["salahtime"].find({}, {"_id": 0}).sort("date", -1).limit(30)
        return list(cur)
    doc = db["salahtime"].find_one({"date": d}, {"_id": 0})
    return doc or {"date": d}


@app.post("/api/salah")
def upsert_salah(item: SalahTime):
    payload = item.model_dump()
    # Ensure date is a JSON-serializable string
    payload["date"] = item.date.isoformat()
    payload["updated_at"] = datetime.utcnow().isoformat()

    if db is None:
        # Fallback: write to local JSON store keyed by date
        store = _read_json(SALAHS_FILE) or {}
        store[payload["date"]] = payload
        _write_json(SALAHS_FILE, store)
        return {"status": "ok", "date": payload["date"], "fallback": True}

    db["salahtime"].update_one(
        {"date": payload["date"]},
        {"$set": {**payload, "updated_at": datetime.utcnow().isoformat()}},
        upsert=True,
    )
    return {"status": "ok", "date": payload["date"]}


# ---------------- Announcements Endpoints ----------------

@app.get("/api/announcements")
def get_active_announcements():
    now = datetime.utcnow()
    if db is None:
        # Fallback: read from JSON and filter like DB would
        items: List[Dict[str, Any]] = _read_json(ANN_FILE) or []
        result = []
        for it in items:
            if not it.get("active", True):
                continue
            start_at = it.get("start_at")
            end_at = it.get("end_at")
            try:
                start_ok = (start_at is None) or (datetime.fromisoformat(start_at.replace("Z", "+00:00")) <= now)
                end_ok = (end_at is None) or (datetime.fromisoformat(end_at.replace("Z", "+00:00")) >= now)
            except Exception:
                start_ok = True
                end_ok = True
            if start_ok and end_ok:
                result.append(it)
        # sort by priority desc
        result.sort(key=lambda x: int(x.get("priority", 1)), reverse=True)
        return result

    filt = {
        "active": True,
        "$and": [
            {"$or": [{"start_at": None}, {"start_at": {"$lte": now}}]},
            {"$or": [{"end_at": None}, {"end_at": {"$gte": now}}]},
        ]
    }
    cur = db["announcement"].find(filt, {"_id": 0}).sort("priority", -1)
    return list(cur)


@app.post("/api/announcements")
def create_announcement(item: Announcement):
    if db is None:
        # Fallback: append to JSON list
        items: List[Dict[str, Any]] = _read_json(ANN_FILE) or []
        data = item.model_dump()
        data["created_at"] = datetime.utcnow().isoformat()
        items.append(data)
        _write_json(ANN_FILE, items)
        return {"status": "ok", "id": len(items), "fallback": True}
    _id = create_document("announcement", item)
    return {"status": "ok", "id": _id}


# ---------------- Assets Endpoints ----------------

@app.get("/api/assets")
def list_assets(limit: int = 20):
    # If DB available, prefer it
    if db is not None:
        cur = db["asset"].find({}, {"_id": 0}).sort("created_at", -1).limit(int(limit))
        return list(cur)

    # Fallback: list files from uploads directory
    items = []
    try:
        entries = sorted(
            (f for f in os.listdir(UPLOAD_DIR) if os.path.isfile(os.path.join(UPLOAD_DIR, f))),
            reverse=True,
        )
        for f in entries[: int(limit)]:
            path = os.path.join(UPLOAD_DIR, f)
            ctype, _ = mimetypes.guess_type(path)
            items.append({
                "filename": f,
                "content_type": ctype or "application/octet-stream",
                "path": f"/uploads/{f}",
            })
    except Exception:
        pass
    return items


# ---------------- File Upload Endpoints ----------------

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file:
        raise HTTPException(status_code=400, detail="No file uploaded")
    filename = file.filename
    # ensure unique filename
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    name, ext = os.path.splitext(filename)
    safe_name = f"{name}_{timestamp}{ext}"
    save_path = os.path.join(UPLOAD_DIR, safe_name)

    content = await file.read()
    with open(save_path, "wb") as f:
        f.write(content)

    url = f"/uploads/{safe_name}"

    # Try to record in DB; if DB not available, still succeed
    try:
        if db is not None:
            meta = Asset(filename=safe_name, content_type=file.content_type or "application/octet-stream", path=url)
            create_document("asset", meta)
    except Exception:
        # Ignore DB errors for upload success
        pass

    return {"status": "ok", "url": url, "content_type": file.content_type or "application/octet-stream"}


# ---------------- AI Sync (Lightweight) ----------------

class AISyncRequest(BaseModel):
    date: str = Field(..., description="YYYY-MM-DD to import")
    commit: bool = Field(True, description="If true, save to store; otherwise preview only")


PRAYER_KEYS = [
    "fajr", "fajr_jamaat", "sunrise",
    "dhuhr", "dhuhr_jamaat",
    "asr", "asr_jamaat",
    "maghrib", "maghrib_jamaat",
    "isha", "isha_jamaat",
]


def _latest_upload() -> Optional[str]:
    try:
        entries = [
            os.path.join(UPLOAD_DIR, f)
            for f in os.listdir(UPLOAD_DIR)
            if os.path.isfile(os.path.join(UPLOAD_DIR, f))
        ]
        if not entries:
            return None
        entries.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return entries[0]
    except Exception:
        return None


def _parse_csv(path: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with open(path, "r", encoding="utf-8") as f:
        sniffer = csv.Sniffer()
        sample = f.read(2048)
        f.seek(0)
        dialect = sniffer.sniff(sample) if sample else csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        for r in reader:
            # normalize keys to lowercase
            norm = { (k or '').strip().lower(): (v or '').strip() for k, v in r.items() }
            rows.append(norm)
    return rows


def _parse_xlsx(path: str) -> List[Dict[str, Any]]:
    """Parse the first worksheet of an XLSX file into list of dict rows with lowercased headers."""
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"XLSX support not available: {str(e)[:120]}")

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell if cell is not None else "" for cell in row])
    if not rows:
        return []
    headers = [str(h).strip().lower() for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        item: Dict[str, Any] = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            val = r[idx] if idx < len(r) else ""
            item[h] = str(val).strip()
        out.append(item)
    return out


def _coerce_times(record: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k in PRAYER_KEYS:
        v = record.get(k)
        if not v:
            continue
        # Normalize common formats like 6:5 to 06:05, and remove am/pm if present
        t = str(v).strip().lower().replace('.', ':').replace(' ', '')
        # Convert times like 615 to 06:15
        if ':' not in t and t.isdigit() and 3 <= len(t) <= 4:
            # e.g., 615 -> 6:15, 0915 -> 09:15
            t = f"{t[:-2]}:{t[-2:]}"
        # remove am/pm after adding colon if needed
        t = t.replace('am', '').replace('pm', '')
        if ':' in t:
            parts = t.split(':')
            try:
                hh = int(parts[0])
                mm = int(parts[1][:2])
                out[k] = f"{hh:02d}:{mm:02d}"
            except Exception:
                continue
    return out


@app.post("/api/sync/ai")
def ai_sync(req: AISyncRequest):
    """
    Lightweight AI-style sync: looks for the most recent uploaded CSV/JSON/XLSX timetable,
    extracts times for the requested date, and optionally commits them.
    Supported formats now: CSV (headers should include columns like fajr, fajr_jamaat, ... and optional date)
    JSON: either an object keyed by date or a list of rows with a 'date' field.
    XLSX: first worksheet with headers in first row.
    """
    target_date = req.date
    # Validate date format
    try:
        datetime.fromisoformat(target_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format; use YYYY-MM-DD")

    path = _latest_upload()
    if not path:
        raise HTTPException(status_code=404, detail="No uploaded files found to sync from")

    _, ext = os.path.splitext(path)
    ext = ext.lower()

    extracted: Dict[str, Any] = {"date": target_date}

    try:
        if ext == ".csv":
            rows = _parse_csv(path)
            # Try to pick row matching date
            match = None
            for r in rows:
                # try keys like 'date' or 'day'
                d = r.get('date') or r.get('day')
                if d:
                    d = str(d).strip()[:10]
                if d == target_date:
                    match = r
                    break
            if match is None and rows:
                # fallback first row
                match = rows[0]
            if match:
                extracted.update(_coerce_times(match))
        elif ext == ".json":
            data = _read_json(path)
            if isinstance(data, dict):
                row = data.get(target_date)
                if isinstance(row, dict):
                    extracted.update(_coerce_times({k.lower(): v for k, v in row.items()}))
            if isinstance(data, list):
                for r in data:
                    if isinstance(r, dict) and (str(r.get('date')) == target_date):
                        extracted.update(_coerce_times({k.lower(): v for k, v in r.items()}))
                        break
        elif ext == ".xlsx":
            rows = _parse_xlsx(path)
            match = None
            for r in rows:
                d = r.get('date') or r.get('day')
                if d:
                    d = str(d).strip()[:10]
                if d == target_date:
                    match = r
                    break
            if match is None and rows:
                match = rows[0]
            if match:
                extracted.update(_coerce_times(match))
        else:
            raise HTTPException(status_code=415, detail=f"Unsupported file type for AI sync: {ext}. Use CSV, JSON, or XLSX.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse source file: {str(e)[:120]}")

    has_any = any(extracted.get(k) for k in PRAYER_KEYS)
    if not has_any:
        raise HTTPException(status_code=422, detail="Could not extract any times from the latest upload. Ensure CSV/JSON/XLSX has columns like fajr, fajr_jamaat, ...")

    if req.commit:
        # Commit to DB or fallback store
        payload = {
            **{k: extracted.get(k) for k in PRAYER_KEYS if extracted.get(k)},
            "date": target_date
        }
        # Use existing upsert logic
        if db is None:
            store = _read_json(SALAHS_FILE) or {}
            existing = store.get(target_date, {})
            existing.update(payload)
            existing["updated_at"] = datetime.utcnow().isoformat()
            store[target_date] = existing
            _write_json(SALAHS_FILE, store)
        else:
            db["salahtime"].update_one(
                {"date": target_date},
                {"$set": {**payload, "updated_at": datetime.utcnow().isoformat()}},
                upsert=True,
            )

    return {
        "status": "ok",
        "source": os.path.basename(path),
        "committed": req.commit,
        "data": {k: extracted.get(k) for k in ["date"] + PRAYER_KEYS if extracted.get(k) or k == "date"}
    }


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
